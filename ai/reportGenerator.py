# ai/reportGenerator.py
import pandas as pd
import numpy as np
import json
import sys
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, PieChart, Reference
import matplotlib.pyplot as plt
import seaborn as sns
from io import BytesIO
import base64

class ReportGenerator:
    def __init__(self):
        self.colors = {
            'header': 'FF4472C4',
            'exact_match': 'FF70AD47',
            'fuzzy_match': 'FFFFC000',
            'ml_match': 'FF5B9BD5',
            'unmatched': 'FFE74C3C',
            'risk_high': 'FFC5504B',
            'risk_medium': 'FFF39C12',
            'risk_low': 'FF27AE60'
        }
    
    def generate_reconciliation_report(self, session_data: dict, output_path: str) -> dict:
        """Generate comprehensive reconciliation report in Excel format"""
        try:
            # Create workbook and worksheets
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            # Create sheets
            summary_sheet = wb.create_sheet("Summary")
            matches_sheet = wb.create_sheet("Matched Transactions")
            unmatched_bank_sheet = wb.create_sheet("Unmatched Bank")
            unmatched_book_sheet = wb.create_sheet("Unmatched Book")
            analytics_sheet = wb.create_sheet("Analytics")
            
            # Generate each section
            self._create_summary_sheet(summary_sheet, session_data)
            self._create_matches_sheet(matches_sheet, session_data.get('matches', []))
            self._create_unmatched_sheets(unmatched_bank_sheet, unmatched_book_sheet, session_data)
            self._create_analytics_sheet(analytics_sheet, session_data)
            
            # Save workbook
            wb.save(output_path)
            
            # Generate summary statistics
            summary_stats = self._calculate_summary_stats(session_data)
            
            return {
                'success': True,
                'report_path': output_path,
                'summary': summary_stats,
                'sheets_created': ['Summary', 'Matched Transactions', 'Unmatched Bank', 'Unmatched Book', 'Analytics']
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': f'Report generation failed: {str(e)}'
            }
    
    def _create_summary_sheet(self, sheet, session_data):
        """Create executive summary sheet"""
        # Title
        sheet['A1'] = 'Bank Reconciliation Report'
        sheet['A1'].font = Font(size=18, bold=True)
        sheet['A1'].fill = PatternFill(start_color=self.colors['header'], end_color=self.colors['header'], fill_type='solid')
        
        # Basic information
        row = 3
        info_data = [
            ('Report Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            ('Session ID:', session_data.get('session_id', 'N/A')),
            ('Processing Date:', session_data.get('processing_date', 'N/A')),
            ('Total Processing Time:', f"{session_data.get('processing_time', 0):.2f} seconds"),
        ]
        
        for label, value in info_data:
            sheet[f'A{row}'] = label
            sheet[f'A{row}'].font = Font(bold=True)
            sheet[f'B{row}'] = value
            row += 1
        
        row += 2
        
        # Transaction counts
        sheet[f'A{row}'] = 'Transaction Summary'
        sheet[f'A{row}'].font = Font(size=14, bold=True)
        sheet[f'A{row}'].fill = PatternFill(start_color=self.colors['header'], end_color=self.colors['header'], fill_type='solid')
        row += 1
        
        # Summary table headers
        headers = ['Category', 'Bank Transactions', 'Book Transactions', 'Matched Pairs', 'Match Rate %']
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color=self.colors['header'], end_color=self.colors['header'], fill_type='solid')
        
        row += 1
        
        # Summary data
        total_bank = session_data.get('bank_transaction_count', 0)
        total_book = session_data.get('book_transaction_count', 0)
        total_matches = session_data.get('total_matches', 0)
        match_rate = (total_matches * 2) / (total_bank + total_book) * 100 if (total_bank + total_book) > 0 else 0
        
        summary_row = ['Total', total_bank, total_book, total_matches, f"{match_rate:.1f}%"]
        for col, value in enumerate(summary_row, 1):
            sheet.cell(row=row, column=col, value=value)
        
        row += 3
        
        # Match type breakdown
        sheet[f'A{row}'] = 'Match Type Breakdown'
        sheet[f'A{row}'].font = Font(size=14, bold=True)
        row += 1
        
        match_types = [
            ('Exact Matches', session_data.get('exact_matches', 0), self.colors['exact_match']),
            ('Fuzzy Matches', session_data.get('fuzzy_matches', 0), self.colors['fuzzy_match']),
            ('ML Matches', session_data.get('ml_matches', 0), self.colors['ml_match']),
            ('Unmatched', (total_bank + total_book) - (total_matches * 2), self.colors['unmatched'])
        ]
        
        for match_type, count, color in match_types:
            sheet[f'A{row}'] = match_type
            sheet[f'B{row}'] = count
            sheet[f'C{row}'] = f"{count/max(total_bank + total_book, 1)*100:.1f}%" if total_bank + total_book > 0 else "0%"
            sheet[f'B{row}'].fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            row += 1
    
    def _create_matches_sheet(self, sheet, matches):
        """Create matched transactions details sheet"""
        headers = [
            'Match ID', 'Bank Transaction ID', 'Book Transaction ID', 'Match Type',
            'Confidence Score', 'Bank Date', 'Book Date', 'Bank Amount', 'Book Amount',
            'Amount Difference', 'Bank Description', 'Book Description', 'Status', 'Risk Factors'
        ]
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color=self.colors['header'], end_color=self.colors['header'], fill_type='solid')
        
        # Add match data
        row = 2
        for i, match in enumerate(matches):
            match_data = [
                i + 1,
                match.get('bank_transaction_id', ''),
                match.get('book_transaction_id', ''),
                match.get('match_type', ''),
                f"{match.get('confidence_score', 0):.3f}",
                match.get('bank_date', ''),
                match.get('book_date', ''),
                f"₹{match.get('bank_amount', 0):.2f}",
                f"₹{match.get('book_amount', 0):.2f}",
                f"₹{abs(match.get('bank_amount', 0) - match.get('book_amount', 0)):.2f}",
                match.get('bank_description', '')[:50],
                match.get('book_description', '')[:50],
                match.get('status', ''),
                ', '.join(match.get('risk_factors', []))
            ]
            
            for col, value in enumerate(match_data, 1):
                cell = sheet.cell(row=row, column=col, value=value)
                
                # Color code by match type
                if match.get('match_type') == 'exact':
                    cell.fill = PatternFill(start_color=self.colors['exact_match'], end_color=self.colors['exact_match'], fill_type='solid')
                elif match.get('match_type') == 'fuzzy':
                    cell.fill = PatternFill(start_color=self.colors['fuzzy_match'], end_color=self.colors['fuzzy_match'], fill_type='solid')
                elif match.get('match_type') == 'ml_suggested':
                    cell.fill = PatternFill(start_color=self.colors['ml_match'], end_color=self.colors['ml_match'], fill_type='solid')
            
            row += 1
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    def _create_unmatched_sheets(self, bank_sheet, book_sheet, session_data):
        """Create sheets for unmatched transactions"""
        # Unmatched bank transactions
        bank_headers = ['Transaction ID', 'Date', 'Amount', 'Description', 'Reference', 'Potential Issues']
        for col, header in enumerate(bank_headers, 1):
            cell = bank_sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color=self.colors['unmatched'], end_color=self.colors['unmatched'], fill_type='solid')
        
        unmatched_bank = session_data.get('unmatched_bank_transactions', [])
        for row, txn in enumerate(unmatched_bank, 2):
            txn_data = [
                txn.get('id', ''),
                txn.get('date', ''),
                f"₹{txn.get('amount', 0):.2f}",
                txn.get('description', '')[:100],
                txn.get('reference', ''),
                self._identify_potential_issues(txn)
            ]
            for col, value in enumerate(txn_data, 1):
                bank_sheet.cell(row=row, column=col, value=value)
        
        # Unmatched book transactions (similar structure)
        book_headers = ['Transaction ID', 'Date', 'Amount', 'Description', 'Reference', 'Potential Issues']
        for col, header in enumerate(book_headers, 1):
            cell = book_sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color=self.colors['unmatched'], end_color=self.colors['unmatched'], fill_type='solid')
        
        unmatched_book = session_data.get('unmatched_book_transactions', [])
        for row, txn in enumerate(unmatched_book, 2):
            txn_data = [
                txn.get('id', ''),
                txn.get('date', ''),
                f"₹{txn.get('amount', 0):.2f}",
                txn.get('description', '')[:100],
                txn.get('reference', ''),
                self._identify_potential_issues(txn)
            ]
            for col, value in enumerate(txn_data, 1):
                book_sheet.cell(row=row, column=col, value=value)
    
    def _create_analytics_sheet(self, sheet, session_data):
        """Create analytics and insights sheet"""
        row = 1
        
        # Performance metrics
        sheet[f'A{row}'] = 'Performance Analytics'
        sheet[f'A{row}'].font = Font(size=16, bold=True)
        row += 2
        
        # Key metrics
        metrics = [
            ('Total Processing Time', f"{session_data.get('processing_time', 0):.2f} seconds"),
            ('Transactions per Second', f"{self._calculate_tps(session_data):.1f}"),
            ('Overall Match Rate', f"{self._calculate_match_rate(session_data):.1f}%"),
            ('Average Confidence Score', f"{self._calculate_avg_confidence(session_data):.3f}"),
            ('Manual Review Required', f"{self._count_manual_reviews(session_data)} transactions"),
        ]
        
        for metric, value in metrics:
            sheet[f'A{row}'] = metric
            sheet[f'A{row}'].font = Font(bold=True)
            sheet[f'B{row}'] = value
            row += 1
        
        row += 2
        
        # Risk analysis
        sheet[f'A{row}'] = 'Risk Analysis'
        sheet[f'A{row}'].font = Font(size=14, bold=True)
        row += 1
        
        risk_summary = self._analyze_risks(session_data)
        for risk_item in risk_summary:
            sheet[f'A{row}'] = risk_item
            row += 1
        
        row += 2
        
        # Recommendations
        sheet[f'A{row}'] = 'Recommendations'
        sheet[f'A{row}'].font = Font(size=14, bold=True)
        row += 1
        
        recommendations = self._generate_recommendations(session_data)
        for rec in recommendations:
            sheet[f'A{row}'] = f"• {rec}"
            row += 1
    
    def _identify_potential_issues(self, transaction):
        """Identify potential issues with unmatched transactions"""
        issues = []
        
        amount = transaction.get('amount', 0)
        description = str(transaction.get('description', '')).upper()
        
        # Check for large amounts
        if abs(amount) > 100000:
            issues.append('Large amount')
        
        # Check for round numbers
        if abs(amount) > 1000 and abs(amount) % 1000 == 0:
            issues.append('Round amount')
        
        # Check for suspicious keywords
        suspicious_words = ['TEST', 'TEMP', 'ERROR', 'MISTAKE']
        if any(word in description for word in suspicious_words):
            issues.append('Suspicious description')
        
        # Check for minimal description
        if len(description.strip()) < 10:
            issues.append('Minimal description')
        
        return ', '.join(issues) if issues else 'None identified'
    
    def _calculate_summary_stats(self, session_data):
        """Calculate summary statistics for the report"""
        total_bank = session_data.get('bank_transaction_count', 0)
        total_book = session_data.get('book_transaction_count', 0)
        total_matches = session_data.get('total_matches', 0)
        
        return {
            'total_transactions': total_bank + total_book,
            'total_matches': total_matches,
            'match_rate_percentage': (total_matches * 2) / (total_bank + total_book) * 100 if (total_bank + total_book) > 0 else 0,
            'unmatched_count': (total_bank + total_book) - (total_matches * 2),
            'processing_time_seconds': session_data.get('processing_time', 0)
        }
    
    def _calculate_tps(self, session_data):
        """Calculate transactions per second"""
        total_txns = session_data.get('bank_transaction_count', 0) + session_data.get('book_transaction_count', 0)
        processing_time = session_data.get('processing_time', 1)
        return total_txns / processing_time if processing_time > 0 else 0
    
    def _calculate_match_rate(self, session_data):
        """Calculate overall match rate"""
        total_bank = session_data.get('bank_transaction_count', 0)
        total_book = session_data.get('book_transaction_count', 0)
        total_matches = session_data.get('total_matches', 0)
        return (total_matches * 2) / (total_bank + total_book) * 100 if (total_bank + total_book) > 0 else 0
    
    def _calculate_avg_confidence(self, session_data):
        """Calculate average confidence score"""
        matches = session_data.get('matches', [])
        if not matches:
            return 0.0
        
        confidence_scores = [m.get('confidence_score', 0) for m in matches]
        return sum(confidence_scores) / len(confidence_scores) if confidence_scores else 0.0
    
    def _count_manual_reviews(self, session_data):
        """Count transactions requiring manual review"""
        matches = session_data.get('matches', [])
        return len([m for m in matches if m.get('manual_review_required', False)])
    
    def _analyze_risks(self, session_data):
        """Analyze risk factors"""
        risks = []
        
        # High-risk matches
        matches = session_data.get('matches', [])
        high_risk_matches = [m for m in matches if m.get('risk_factors')]
        if high_risk_matches:
            risks.append(f"• {len(high_risk_matches)} matches have risk factors requiring attention")
        
        # Low confidence matches
        low_confidence = [m for m in matches if m.get('confidence_score', 1) < 0.7]
        if low_confidence:
            risks.append(f"• {len(low_confidence)} matches have low confidence scores")
        
        # Large unmatched amounts
        unmatched_bank = session_data.get('unmatched_bank_transactions', [])
        large_unmatched = [t for t in unmatched_bank if abs(t.get('amount', 0)) > 50000]
        if large_unmatched:
            risks.append(f"• {len(large_unmatched)} large amount transactions remain unmatched")
        
        return risks if risks else ["• No significant risks identified"]
    
    def _generate_recommendations(self, session_data):
        """Generate actionable recommendations"""
        recommendations = []
        
        match_rate = self._calculate_match_rate(session_data)
        
        if match_rate < 80:
            recommendations.append("Consider adjusting matching tolerance settings to improve match rate")
        
        manual_reviews = self._count_manual_reviews(session_data)
        total_matches = session_data.get('total_matches', 1)
        if manual_reviews / total_matches > 0.3:
            recommendations.append("High manual review rate - consider model retraining")
        
        avg_confidence = self._calculate_avg_confidence(session_data)
        if avg_confidence < 0.8:
            recommendations.append("Low average confidence - additional training data may improve accuracy")
        
        unmatched_count = len(session_data.get('unmatched_bank_transactions', []))
        if unmatched_count > 50:
            recommendations.append("High number of unmatched transactions - review data quality and formats")
        
        return recommendations if recommendations else ["System performing well - continue regular monitoring"]

def main():
    try:
        # Read input from stdin
        input_data = json.loads(sys.stdin.read())
        
        generator = ReportGenerator()
        
        # Extract parameters
        session_data = input_data.get('session_data', {})
        output_path = input_data.get('output_path', 'reconciliation_report.xlsx')
        report_type = input_data.get('report_type', 'summary')
        
        # Generate report
        result = generator.generate_reconciliation_report(session_data, output_path)
        
        print(json.dumps(result, default=str))
        
    except Exception as e:
        error_result = {
            'success': False,
            'error': str(e)
        }
        print(json.dumps(error_result))

if __name__ == "__main__":
    main()